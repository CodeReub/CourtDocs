import tkinter
import tkinter.ttk
from tkinter import filedialog
from openpyxl import load_workbook
from datetime import date, datetime
import re



# main, establishes the Tkinter interface window
class GUI:
    def __init__(self):
        # instantiate main window
        window = tkinter.Tk()
        window.title('Court Docs')
        window.geometry('540x450')

        # initialize the selector variable
        window.radio_var = tkinter.IntVar()

        # add components to main window
        GUI.main_menu(window)
        GUI.greeting(window)
        GUI.selector(window)

        window.mainloop()
        
    # sample menus (no functionality yet, for future development only
    def main_menu(window):
        menu = tkinter.Menu(window)
        # File menu
        file_menu = tkinter.Menu(menu)
        file_menu.add_command(label='New')
        file_menu.add_command(label='Open')
        file_menu.add_command(label='Exit')
        file_menu.add_command(label='Quit')
        menu.add_cascade(label='File', menu=file_menu)
        # Project menu
        project_menu = tkinter.Menu(menu)
        project_menu.add_command(label='New')
        project_menu.add_command(label='Open')
        project_menu.add_command(label='Exit')
        project_menu.add_command(label='Quit')
        menu.add_cascade(label='Project', menu=project_menu)
        
        window.config(menu=menu)

    # basic greeting label
    def greeting(window):
        greeting = tkinter.ttk.Label(window, text="Welcome to Court Docs Converter",
                                 font=("Times", 20), padding=(30, 6))
        greeting.grid(column=0, row=0)
        copyr = tkinter.ttk.Label(window, text="Copyright 2018, Reuben Mees",
                                  font=("Times", 8), padding=(50, 2))
        copyr.grid(column=0, row=1)

    ### the Radiobutton program selector
    def selector(window):

        # simply variable name in this function only
        x=window.radio_var.get()

        # define the radio buttons
        rb1 = tkinter.Radiobutton(window, text='Municipal Court Criminal',
                                  variable=window.radio_var, value=1)#, command=window.radio_var.set(1))
        rb2 = tkinter.Radiobutton(window, text='Municipal Court Bonds',
                                  variable=window.radio_var, value=2)#, command=window.radio_var.set(2))
        rb3 = tkinter.Radiobutton(window, text='Municipal Court Civil',
                                  variable=window.radio_var, value=3)#, command=window.radio_var.set(3))
        rb4 = tkinter.Radiobutton(window, text='Common Pleas Civil',
                                  variable=window.radio_var, value=4)#, command=window.radio_var.set(4))
        rb5 = tkinter.Radiobutton(window, text='Real Estate Transfers',
                                  variable=window.radio_var, value=5)#, command=window.radio_var.set(5))
        rb6 = tkinter.Radiobutton(window, text='Marriage Licenses',
                                  variable=window.radio_var, value=6)#, command=window.radio_var.set(6))
        rb7 = tkinter.Radiobutton(window, text='Fair Premiums',
                                  variable=window.radio_var, value=7)#, command=window.radio_var.set(7))
        contbutton = tkinter.Button(window, text="Continue", command=lambda:GUI.pickprog(window.radio_var.get()))
#        exitbutton = tkinter.Button(window, text="Continue",
 #                                   command=window.destroy())


        # pack the buttons
        rb1.grid(column=0, row=3)
        rb2.grid(column=0, row=4)
        rb3.grid(column=0, row=5)
        rb4.grid(column=0, row=6)
        rb5.grid(column=0, row=7)
        rb6.grid(column=0, row=8)
        rb7.grid(column=0, row=9)
        contbutton.grid(column=0, row=12)




    # function to select type of file to convert
    #   called from end of Radiobutton -- selector() function
    def pickprog(opt):
        print(opt)
       # opt = window.radio_var.get()
        if opt==1:
            MCT()
        elif opt==2:
            MCT_bonds()
        elif opt==3:
            MCT_civil()
        elif opt==4:
            CCP_civil()
        elif opt==5:
            transfers()
        elif opt==6:
            marriages()
        elif opt==7:
            fair_OC()
    #    else:
    #        ??

# file dialog for opening files
def load_file():
    return tkinter.filedialog.askopenfilename()



# function for setting date to AP style, used im marriages.py, CCP_civil, transfers
def AP_Style_date(start_date):
    new_date=start_date.strftime("%b. %d")
    
    #eliminates 0 from days 01-09
    if new_date[5]=="0": 
        new_date=new_date[0:5]+new_date[6:]

    #formats months to AP style
    final_date=new_date
    if new_date[0:4]=="Mar.":
        final_date="March"+new_date[4:]
    elif new_date[0:4]=="Apr.":
        final_date="April"+new_date[4:]
    elif new_date[0:4]=="May.":
        final_date="May"+new_date[4:]
    elif new_date[0:4]=="Jun.":
        final_date="June"+new_date[4:]
    elif new_date[0:4]=="Jul.":
        final_date="July"+new_date[4:]
    elif new_date[0:4]=="Sep.":
        final_date="Sept."+new_date[4:]

    return final_date


# name format function for bonds.py, MCT.py, may be usable for marriages.py, CCP_civil.py
def name_field1(nm):
    if nm==None:
        name=" "
    try:
        name=nm.title()
    except AttributeError:
        name=" "
    if name==None:
        name=" "
    name=name.split(", ")

    #define last name
    last_name=name[0]
    if last_name[0:2]=="Mc":
        last_name=last_name[0:2]+last_name[2].upper()+last_name[3:len(last_name)]
    if last_name[0:5]=="Levan":
        last_name=last_name[0:2]+last_name[2].upper()+last_name[3:len(last_name)]

    #define first name & add periods
    try:
        first_name=name[1]
    except IndexError:
        first_name=""
    try:
        if first_name[len(first_name)-2]==" ":  #add a period after a middle initial
            first_name=first_name+"."
    except IndexError:
        first_name=""
    try:
        if first_name[1]==" ":   #add a period after first initial
            first_name=first_name[0]+"."+first_name[1:]
    except IndexError:
        first_name=""

    #testing for & setting the suffix
    try:
        suffix_name=name[2]
    except IndexError:
        suffix_name=""
    if suffix_name=="Jr":
        suffix_name="Jr."
    elif suffix_name=="Sr":
        suffix_name="Sr."
    elif suffix_name=="Ii":
        suffix_name="II"
    elif suffix_name=="Iii":
        suffix_name="III"
    elif suffix_name=="Iv":
        suffix_name="IV"

    #returning the new name field
    name=(first_name+" "+last_name+" "+suffix_name)
    try:
        while name[len(name)-1]==" ":
            name=name[0:len(name)-1]
    except IndexError:
        name=""
    return name

# function to format a name for marriages.py, CCP_civil.py (may not be needed (see above))
def name_field(nm):
    name=nm.split(", ")
    suffix_name=""

    #define last name
    last_name=name[0]
    if last_name[0:2]=="Mc":
        last_name=last_name[0:2]+last_name[2].upper()+last_name[3:len(last_name)]
    if last_name[0:5]=="Levan":
        last_name=last_name[0:2]+last_name[2].upper()+last_name[3:len(last_name)]

    #define first name & add periods
    try:
        if name[1]=="Jr" or name[1]=="Sr" or name[1]=="I" or name[1]=="Ii" or name[1]=="Iii" or name[1]=="Iv":
            suffix_name=name[1]
            first_name=name[2]
        else:
            first_name=name[1]
    except IndexError:
        first_name=""
        
    #add a period after a middle initial
    try:
        if first_name[len(first_name)-2]==" ":
            first_name=first_name+"."
    except IndexError:
        first_name=""
        
    #add a period after first initial
    try:
        if first_name[1]==" ":
            first_name=first_name[0]+"."+first_name[1:]
    except IndexError:
        first_name=""

    #testing for & setting the suffix
    if suffix_name=="Jr":
        suffix_name="Jr."
    elif suffix_name=="Sr":
        suffix_name="Sr."
    elif suffix_name=="Ii":
        suffix_name="II"
    elif suffix_name=="Iii":
        suffix_name="III"
    elif suffix_name=="Iv":
        suffix_name="IV"

    #returning the new name field
    name=(first_name+" "+last_name+" "+suffix_name)
    while name[len(name)-1]==" ":
        name=name[0:len(name)-1]
    return name

# function to calculate age from date of birth for marriages.py
def calculate_age(born):
    today=datetime.today()
    try:
        birthday=born.replace(year=today.year)
    except ValueError: #exception for Feb. 29  birthday
        birthday=born.replace(year=today.year, month=born.month+1, day=1)
    if birthday>today:
        return today.year-born.year-1
    else:
        return today.year-born.year




# City, State function for bonds.py, MCT.py
def location(city_cell):
    try:
        city_state=city_cell.title()
    except AttributeError:
        city_state=""
    city_state=city_state.split(", ")

    #define city
    city=city_state[0]
    if city=="Degraff":
        city="DeGraff"

    #define state
    try:
        state=city_state[1]
    except IndexError:
        state=""
    if state=="Oh":
        state=""
    elif state=="Ga" or state=="Ky" or state=="La" or state=="Md" or state=="Mo" \
         or state=="Pa" or state=="Vt" or state=="Va":
        state=state+"."
    elif state=="Nh" or state=="Nj" or state=="Nm" or state=="Ny" or state=="Nc" \
         or state=="Nd" or state=="Ri" or state=="Sc" or state=="Sd":
        state=state[0]+"."+state[1].upper()+"."
    elif state=="Al":
        state="Ala."
    elif state=="Ak":
        state="Alaska"
    elif state=="Az":
        state="Ariz."
    elif state=="Ar":
        state="Ark."
    elif state=="Ca":
        state="Calif."
    elif state=="Co":
        state="Colo."
    elif state=="Ct":
        state="Conn."
    elif state=="De":
        state="Del."
    elif state=="Fl":
        state="Fla."
    elif state=="Hi":
        state="Hawaii"
    elif state=="Id":
        state="Idaho"
    elif state=="Il":
        state="Ill."
    elif state=="Ia":
        state="Iowa"
    elif state=="In":
        state="Ind."
    elif state=="Ks":
        state="Kan."
    elif state=="Me":
        state="Maine"
    elif state=="Ma":
        state="Mass."
    elif state=="Mi":
        state="Mich."
    elif state=="Mn":
        state="Minn."
    elif state=="Ms":
        state="Miss."
    elif state=="Mt":
        state="Mont."
    elif state=="Ne":
        state="Neb."
    elif state=="Nv":
        state="Nev."
    elif state=="Ok":
        state="Okla."
    elif state=="Or":
        state="Ore."
    elif state=="Tn":
        state="Tenn."
    elif state=="Tx":
        state="Texas"
    elif state=="Ut":
        state="Utah"
    elif state=="Wa":
        state="Wash."
    elif state=="Wv":
        state="W.Va."
    elif state=="Wi":
        state="Wis."
    elif state=="Wy":
        state="Wyo."
    
    #combine city, state
    if city=="" or city==" ":
        city_and_state="at large"
    elif state=="":
        city_and_state=city
    else:
        city_and_state=city+", "+state
    return city_and_state


# charge & fine function for bonds.py
def charge_fine_bonds(charge, fine, test_name, tempi, sheet):
    try:
        charge=charge.lower()
    except AttributeError:
        charge=""
    if charge=="use/possess marihuana parapher":
        charge="use/possess marijuana parapher"
    elif charge=="possession of marihuana":
        charge="possession of marijuana"
    elif charge=="fty- while turning left":
        charge="failure to yield while turning left"
    elif charge=="fyr of way":
        charge="failure to yield right of way"
    elif charge=="fail/yield at stop sign":
        charge="fail to yield at stop sign"
    
    #fine amount
    fine=str(fine)
    try:
        if fine[len(fine)-2]==".":  #add a zero in cents place if needed
            fine=fine+"0"
    except IndexError:
        fine=""

    #checking if charge continues to lower line
    end=False
    charge_add=""
    while test_name==None and end!=True:
        charge_add_number="E"+str(tempi)
        try:
            charge_add=sheet[charge_add_number].value.lower()
        except AttributeError:
            end=True
        charge=charge+charge_add
        tempi+=1
        tn_number="A"+str(tempi)
        test_name=sheet[tn_number].value

    if fine=="0":
        chargefine=charge+", no fine"
    else:
        chargefine=charge+", $"+fine+" fine"
    return chargefine, tempi, test_name

# charge & fine function for MCT.py
def charge_fine(ch, fi, st, co, jt, jc, tn, tempi, sh):
    try:
        charge=ch.lower()
    except AttributeError:
        charge=""
    if charge=="use/possess marihuana parapher":
        charge="possession of marijuana parapher"
    elif charge=="possession of marihuana":
        charge="possession of marijuana"
    elif charge=="fty- while turning left":
        charge="failure to yield while turning left"
    elif charge=="fyr of way":
        charge="failure to yield right of way"
    elif charge=="fail/yield at stop sign":
        charge="fail to yield at stop sign"
    elif charge=="f/yield right of way":
        charge="failure to yield right of way"
    elif charge=="dus/lic. forf. or child suppor":
        charge="driving under suspension/license forfeiture or child suppor"
    elif charge=="driving under fra suspension":
        charge="driving under FRA suspension"
    elif charge=="driving under ovi suspension":
        charge="driving under OVI suspension"
    elif charge=="ovi/refusal":
        charge="OVI refusal"
    elif charge=="operate veh. under influence":
        charge="OVI"
    elif charge=="cons/possess alcohol underage":
        charge="underage consumption or possession of alcohol"
    elif charge=="f/to comply with officer":
        charge="failure to comply with an officer"
    elif charge=="f/have safety equip. on board":
        charge="failure to have safety equipment on board"
    elif charge=="unauth. use/fict. plates/tags":
        charge="unauthorized use/fictitious plates or tags"
    elif charge=="obed. to ped./traffic devices":
        charge="obedience to pedestrian traffic devices"
    elif charge=="u\" turns restricted\"":
        charge="making a restricted U-turn"
    elif charge=="physical control w/intoxicated":
        charge="physical control of a motor vehicle while intoxicated"
    elif charge=="driving under fra/judgment suspension":
        charge="driving under FRA suspension"
    elif charge=="driving w/o valid license":
        charge="driving without a valid license"
    elif charge=="fail/display valid plate/reg":
        charge="failure to display valid plate or registration"
    elif charge=="f/display license plate":
        charge="failure to display license plate"
    elif charge=="no tail light/lic. plate light":
        charge="no tail light or license plate light"
    elif charge=="f/complete boat safety course":
        charge="failure to complete boat safety course"
    elif charge=="driving w/intoxicated":
        charge="OVI"
    elif charge=="f/to raze or repair structure":
        charge="failure to raze or repair structure"
    elif charge=="permit holder w/o lic. driver":
        charge="driving with a temporary permit"
    elif charge=="traffic control device/sign":
        charge="failure to obey traffic control device or sign"
    elif charge=="operation w/o pfd prohibited":
        charge="driving without insurance"
    elif charge=="f/drive in marked lanes":
        charge="failure to drive in marked lanes"        
    elif charge=="f/display license":
        charge="failure to display license"        
    elif charge=="consume alcohol in m.v.":
        charge="consuming alcoholin a motor vehicle"        
    elif charge=="operating veh. under influence":
        charge="OVI"

    #status/plea
    status=""
    if str(st)=="G" or str(st)=="WG" or str(st)=="TCOP" or str(st)=="COP" or str(st)=="COPP" or str(st)=="PTR" or str(st)=="GRT" or str(st)=="GMS" or str(st)=="RMS" or str(st)=="GT":
        status="guilty"
    elif str(st)=="NGA" or str(st)=="NG":
        status="not guilty"
    elif str(st)=="DIS" or str(st)=="DISM" or str(st)=="DSMC" or str(st)=="TD" or str(st)=="TDC" or str(st)=="DMS" or str(st)=="PTD":
        status="dismissed"

    #fine amount
    fine=str(fi)
    try:
        if fine[len(fine)-2]==".":  #add a zero in cents place if needed
            fine=fine+"0"
    except IndexError:
        fine=""
    if fine=="None":
        fine=""
        
    try:
        if fine[len(fine)-3]=="." and len(fine)>6:
            fine=fine[:len(fine)-6]+","+fine[len(fine)-6:]
        elif len(fine)>3 and fine[len(fine)-3]!=".":
            fine=fine[:len(fine)-3]+","+fine[len(fine)-3:]

    except IndexError:
        pass

    #miscellaneous cost amount
    cost=str(co)
    try:
        if cost[len(cost)-2]==".":  #add a zero in cents place if needed
            cost=cost+"0"
    except IndexError:
        cost=""

    #checking if charge continues to lower line
    while tn==None and tempi!=sh.max_row:
        try:
            charge+=sh["H"+str(tempi)].value.lower()
        except AttributeError:
            pass
        tempi+=1
        tn=sh["B"+str(tempi)].value

    if fine=="" or fine=="0":
        chargefine=charge+", "+status
    else:
        chargefine=charge+", "+status+", $"+fine+" fine"
    if cost!=None and cost!="0" and cost!="None":
        chargefine=chargefine+", $"+cost+" costs"
    if jt!=0 and jt!=None:
        chargefine=chargefine+", "+str(jt)+" days in jail"
    if jc!=0 and jc!=None:
        chargefine=chargefine+", "+str(jc)+" days suspended"
    return chargefine, tempi, tn


# functions for formatting party names in MCT_civil.py
def middle(name):
    parts=name.split(" ")
    newname=""
    for part in parts:
        if len(part)==1 and part.isalpha():
            part=part+"."
        if newname=="":
            newname=part
        else:
            newname+=" "+part

    newname.replace("  "," ")

    return(newname)

# find & replace function for MCT_civil, CCP_civil files
def findreplace(info):
    info=info.replace(r"  ", " ")
    info=info.replace(r"  ", " ")
    info=info.replace(r"  ", " ")
    info=info.replace(r" Llc"," LLC")
    info=info.replace(r" Ll"," LLC")
    info=info.replace(r" L.L.C."," LLC")
    info=info.replace(r", LLC"," LLC")
    info=info.replace(r" Incorportated"," Inc.")
    info=info.replace(r", Inc"," Inc.")
    info=info.replace(r" Inc: ",r" Inc.: ")
    info=info.replace(r" Inc "," Inc. ")
    info=info.replace(r" Corporation"," Corp.")
    info=info.replace(r", Corp"," Corp.")
    info=info.replace(r" Corp: ",r" Corp.: ")
    info=info.replace(r" : ",": ")
    info=info.replace(r" Iii"," III")
    info=info.replace(r" Ii"," II")
    info=info.replace(r" Jr"," Jr.")
    info=info.replace(r"..",".")
    info=info.replace(r" Sr"," Sr.")
    info=info.replace(r"Lvnv","LVNV")
    info=info.replace(r"Sac ","SAC ")
    info=info.replace(r" (Usa) Na","")
    info=info.replace(r" (Usa), Na","")
    info=info.replace(r" (Usa) N.A.","")
    info=info.replace(r" (Usa), N.A.","")
    info=info.replace(r" N.A.","")
    info=info.replace(r" Na "," ")
    info=info.replace(r"( Usa)","")
    info=info.replace(r" Of "," of ")
    info=info.replace(r"Spv I.","SPV I")
    info=info.replace(r"1St ","1st ")
    info=info.replace(r"Jh ","JH ")
    info=info.replace(r"Ih ","IH ")
    info=info.replace(r"Pca ","PCA ")
    info=info.replace(r"Cu ","CU ")
    info=info.replace(r".00","")
    info=info.replace(r" And "," and ")
    info=info.replace(r"Jpmorgan","JPMorgan")
    info=info.replace(r"Unknown Spouse","")
    info=info.replace(r"Unk Spouse If Any","")
    info=info.replace(r"Us Bank","U.S. Bank")
    info=info.replace(r"'S",r"'s")
    info=info.replace(r"Jane Doe, ","")
    info=info.replace(r"John Doe, ","")
    info=info.replace(r"Doe, ","")
    info=info.replace(r"Fsb ","FSB ")
    info=info.replace(r"Ntl Collegiate Stndt Ln Trst","National Collegiate Student Loan Trust")
    info=info.replace(r"  "," ")
    info=info.replace(r", , ",r", ")
    info=info.replace(r" : ",r": ")
    
    return(info)


# function to find the property location from its tax code - transfers.py
def tax_code(tloc):
    if tloc=="01" or tloc=="02" or tloc=="03":
        floc="Bloomfield Township"
    elif tloc=="04" or tloc=="05":
        floc="Bokescreek Township"
    elif tloc=="06" or tloc=="33":
        floc="West Mansfield"
    elif tloc=="07":
        floc="Ridgeway"
    elif tloc=="08" or tloc=="09" or tloc=="10":
        floc="Harrison Township"
    elif tloc=="11" or tloc=="17" or tloc=="55" or tloc=="57" or tloc=="58" or tloc=="59" or tloc=="60" or tloc=="61":
        floc="Bellefontaine"
    elif tloc=="12":
        floc="Jefferson Township"
    elif tloc=="13" or tloc=="31":
        floc="Valley Hi"
    elif tloc=="14":
        floc="Zanesfield"
    elif tloc=="15" or tloc=="16":
        floc="Lake Township"
    elif tloc=="18" or tloc=="19":
        floc="Liberty Township"
    elif tloc=="20" or tloc=="56":
        floc="West Liberty"
    elif tloc=="21" or tloc=="22" or tloc=="23":
        floc="McArthur Township"
    elif tloc=="24":
        floc="Huntsville"
    elif tloc=="25":
        floc="Miami Township"
    elif tloc=="26" or tloc=="35":
        floc="DeGraff"
    elif tloc=="27":
        floc="Quincy"
    elif tloc=="28" or tloc=="29" or tloc=="30":
        floc="Monroe Township"
    elif tloc=="32":
        floc="Perry Township"
    elif tloc=="34":
        floc="Pleasant Township"
    elif tloc=="36" or tloc=="37" or tloc=="38":
        floc="Richland Township"
    elif tloc=="39":
        floc="Belle Center"
    elif tloc=="40" or tloc=="41":
        floc="Rushcreek Township"
    elif tloc=="42":
        floc="Rushsylvania"
    elif tloc=="43" or tloc=="44" or tloc=="45" or tloc=="46":
        floc="Stokes Township"
    elif tloc=="47":
        floc="Lakeview"
    elif tloc=="48" or tloc=="49" or tloc=="50":
        floc="Union Township"
    elif tloc=="51":
        floc="Washington Township"
    elif tloc=="52":
        floc="Russells Point"
    elif tloc=="53" or tloc=="54":
        floc="Zane Township"
    return floc

# function that modifies the grantor and grantee fields, transfers.py
def grantor_grantee(gr):
    pattern=r"^[A-Z] " #check for first initial only
    match=re.search(pattern, gr)
    if match:
        gr=gr[0]+"."+gr[1:]
    pattern=r" [A-Z] " #check for middle initials
    match=re.search(pattern, gr)
    while match:
        gr=gr[0:match.start()+2]+". "+gr[match.end():]
        match=re.search(pattern, gr)
    pattern=r" Mc[a-z]" #Last names that start with Mc
    match=re.search(pattern, gr)
    if match:
        gr=gr[0:match.start()+3]+gr[match.start()+3].capitalize()+gr[match.end():]
    pattern=r"-Mc[a-z]" #Last names that start with -Mc
    match=re.search(pattern, gr)
    if match:
        gr=gr[0:match.start()+3]+gr[match.start()+3].capitalize()+gr[match.end():]

    # standard find and replace statements
    gr=gr.replace("  ", " ")
    gr=gr.replace("  ", " ")
    gr=gr.replace("  ", " ")
    gr=gr.replace("  ", " ")
    gr=gr.replace("  ", " ")
    gr=gr.replace(" Etal", "")
    gr=gr.replace(" Etux", "")
    gr=gr.replace(" Etvir", "")
    c=2   # find & rplace any partial interest up to 1/10
    while c<=10:
        gr=gr.replace(" 1/"+str(c)+" Int", "")
        c+=1
    gr=gr.replace(" & ", " and ")
    gr=gr.replace(" Trustee", ", trustee,")
    gr=gr.replace(" Co ", " Co. ")
    gr=gr.replace(" Co,", " Co.,")
    gr=gr.replace(" Jr", " Jr.")
    gr=gr.replace(" Sr", " Sr.")
    gr=gr.replace(" Ii", " II")
    gr=gr.replace(" Iii", " III")
    gr=gr.replace(" Llc", " LLC")
    gr=gr.replace(" Inc", " Inc.")
    gr=gr.replace("Mr ", "")
    gr=gr.replace("Mrs  ", "")
    gr=gr.replace("Ms ", "")
    
    return(gr)

def check_line(lin):
    while "  " in lin:
        lin=lin.replace("  ", " ")
    while ",," in lin:
        lin=lin.replace(",,", ",")

    return(lin)



# bonds.py --- body of the original program
def MCT_bonds():
    #load the spreadsheet file, create the text file
    fn=load_file()
    wb=load_workbook(filename=fn)
    sh=wb.active
    last_row=sh.max_row
    textfile=open("bonds.txt","w")
    textfile.write("Bonds forfeited\nThe following bonds were recently forfeited in Bellefontaine Municipal Court:\n")

    #main loop
    i=2
    while i<=last_row:
        
        #identifying name cell and calling function
        name_cell_number="A"+str(i)
        name_cell=sh[name_cell_number].value
        Name=name_field1(name_cell).title()

        #identifying city, state cell and calling function
        city_cell_number="D"+str(i)
        city_cell=sh[city_cell_number].value
        CityState=location(city_cell)

        #identifying charge and fine cells and calling function
        ##also checks for multiple charges
        test_name=name_cell
        ChargeFine=""
        while test_name==name_cell:
            charge_cell_number="E"+str(i)
            charge_cell=sh[charge_cell_number].value
            fine_cell_number="F"+str(i)
            fine_cell=sh[fine_cell_number].value
            i+=1
            test_name_number="A"+str(i)
            test_name=sh[test_name_number].value
            Chargefine, i, test_name=charge_fine_bonds(charge_cell, fine_cell,
                                                       test_name, i, sh)
            if ChargeFine=="":
                ChargeFine=Chargefine
            else:
                ChargeFine=ChargeFine+"; "+Chargefine


        
        textfile.write(Name+", of "+CityState+": "+ChargeFine+"\n")
           
        
    textfile.close()


# MCT.py -- body of the original program
def MCT():
    #load the spreadsheet file, create the text file
    fn=load_file()
    wb=load_workbook(filename=fn)
    sh=wb.active
    textfile=open("MCT.txt","w")
    textfile.write("Municipal Court\nThe following cases were recently adjudged in Bellefontaine Municipal Court:\n")

    #main loop
    last_row=sh.max_row
    print(last_row+1) # debug number only
    i=2
    while i<=last_row:
        
        #identifying name cell and calling function
        name_cell=sh["B"+str(i)].value
        Name=name_field1(name_cell)

        #identifying city, state cell and calling function
        city_cell=sh["E"+str(i)].value
        CityState=location(city_cell)

        #identifying charge and fine cells and calling function
        ##also checks for multiple charges
        test_name=name_cell
        ChargeFine=""
        while test_name==name_cell:
            charge_cell=sh["H"+str(i)].value
            fine_cell=sh["P"+str(i)].value
            status_cell=sh["L"+str(i)].value
            cost_cell=sh["Q"+str(i)].value
            jailtime_cell=sh["R"+str(i)].value
            jailcredit_cell=sh["S"+str(i)].value
            i += 1
            test_name=sh["B"+str(i)].value
            Chargefine, i, test_name=charge_fine(charge_cell, fine_cell, status_cell,
                                                 cost_cell, jailtime_cell, jailcredit_cell,
                                                 test_name, i, sh)
            if ChargeFine=="":
                ChargeFine=Chargefine
            else:
                ChargeFine=ChargeFine+"; "+Chargefine


                
        print(Name) # for debug output only
    
        textfile.write(Name+", of "+CityState+": "+ChargeFine+"\n")
            
    textfile.close()

# marriage.py -- body of the original program
def marriages():
    #open files
    fn=load_file()
    readfile=open(fn, "r")
    
    writefile=open("marry.txt", "w")
    writefile.write("Marriages\n")
    writefile.write("The following couples recently filed for marriage licenses in Logan County Family Court:\n")

  #main loop
    party1=""
    party2=""
    oldcase=""
    filedate=""
    age1=""
    age2=""
    for line in readfile:
        data=line.split("\t")
        if data[0]==" ":   #remove extra spaces from start of case number
            case=data[0][1:12]
        else:
            case=data[0][0:11]
        if case!=oldcase:   #instructions for when case number changes
            if oldcase!="": #write to file only if it's not the first case
                writefile.write(filedate+": "+party1+", "+age1+", and "+party2+", "+age2+"\n")
            oldcase=case
            fdate=datetime.strptime(data[2][0:10], "%m/%d/%Y")
            filedate=AP_Style_date(fdate)
            print(filedate)
        party=data[5][0:11]  #party is party type (i.e App 1 or 2), not the party name
        party_name=str(data[4].title()) #this is the party name
        while party_name[len(party_name)-1]==" ":
            party_name=party_name[0:len(party_name)-1]
        while party_name[0]==" ":
            party_name=party_name[1:len(party_name)]    
        party_name=name_field(party_name) #this runs the name_field() function
        print(party_name, party)  ###this is an unnecesary line for debugging purposes only###
        if party=="Applicant 1":
            party1=party_name
            DOB=datetime.strptime(data[6][0:10], "%m/%d/%Y")
            age1=str(calculate_age(DOB)) #this runs the calculate_age() function
        elif party=="Applicant 2":
            party2=party_name
            DOB=datetime.strptime(data[6][0:10], "%m/%d/%Y")
            age2=str(calculate_age(DOB)) #this runs the calculate_age() function

    #write the final entry and close
    writefile.write(filedate+": "+party1+", "+age1+", and "+party2+", "+age2+"\n")
    readfile.close()
    writefile.close()


# MCT_civil.py --- original body
def MCT_civil():
    #intro-creating output file
    print("Make sure there are up to 7 .xlsx files in this directory with the names:\n")
    print("JUDG, DEFJ, DSLP, DSPL, DSJU, CLPD and JUPD")
    non=str(input("Hit Enter to continue"))
    writefile=open("MCT_civil.txt", "w")
    writefile.write("Civil judgments\nThe following judgments were recently awarded in Bellefontaine Municipal Court:\r")

    #load the spreadsheet workbook
    filename='JUDG.xlsx'
    while filename=='JUDG.xlsx' or filename=='DEFJ.xlsx':
        try:
            wb=load_workbook(filename)
            sh=wb.active
            last_row=sh.max_row
            oldplaint=""
            line=""
            i=3
            while i<=last_row:
                #reading four main cells
                plaint_num="C"+str(i)
                plaintiff=sh[plaint_num].value
                plaintiff=plaintiff.title()
                plaintiff=middle(plaintiff)
                def1_num="D"+str(i)
                def1=sh[def1_num].value
                def1=def1.title()
                def1=middle(def1)
                def2_num="E"+str(i)
                try:     #check def2 has a value
                    def2=sh[def2_num].value
                    def2=def2.title()
                    def2=middle(def2)
                except AttributeError:
                    def2=""
                judg_num="F"+str(i)
                try:     #check judgment has value & format currency
                    judg=float(sh[judg_num].value)
                    judg="${:,.2f}".format(judg)
                except TypeError:
                    judg=""
                if def2=="":    #single defendant
                    defs=def1
                else:           #2 defendants
                    defs=def1+" and "+def2
                    
                if judg!="" and judg!="$0.00":   #does not print non judgments
                    if plaintiff==oldplaint:     #single plaintiff, multiple cases
                        line=line+"; "+judg+" from "+defs
                    else:     #print line to file & move on to new plaintiff
                        line=findreplace(line)
                        writefile.write(line+"\n")
     #                   print(line)    #this is for testing only***
                        line=plaintiff+": "+judg+" from "+defs
                if i==last_row:    #print last entry if about to jump out of loop
                    line=findreplace(line)
                    writefile.write(line+"\n")
     #               print(line)    #this is for testing only***
                oldplaint=plaintiff
                i+=1
            wb.save(filename)
        except FileNotFoundError:
            print("No file: "+filename)
        if filename=="JUDG.xlsx":
            filename="DEFJ.xlsx"
        else:
            filename=""

    #starting dismissals (5 separate case codes)
    writefile.write("Dismissals\nThe following civil claims were recently dismissed in Bellefontaine Municipal Court:\r")

    #load the spreadsheet workbook
    filename='DSLP.xlsx'
    while filename=='DSLP.xlsx' or filename=='DSPL.xlsx' or filename=='DSJU.xlsx' or filename=='CLPD.xlsx' or filename=='JUPD.xlsx':
        try:
            wb=load_workbook(filename)
            sh=wb.active
            last_row=sh.max_row
            oldplaint=""
            line=""
            i=3
            while i<=last_row:
                #reading four main cells
                plaint_num="C"+str(i)
                plaintiff=sh[plaint_num].value
                plaintiff=plaintiff.title()
                plaintiff=middle(plaintiff)
                def1_num="D"+str(i)
                def1=sh[def1_num].value
                def1=def1.title()
                def1=middle(def1)
                def2_num="E"+str(i)
                try:     #check def2 has a value
                    def2=sh[def2_num].value
                    def2=def2.title()
                    def2=middle(def2)
                except AttributeError:
                    def2=""
                if def2=="":    #single defendant
                    defs=def1
                else:           #2 defendants
                    defs=def1+" and "+def2
                if plaintiff==oldplaint:     #single plaintiff, multiple cases
                    line=line+"; vs. "+defs
                else:     #print line to file & move on to new plaintiff
                    line=findreplace(line)
                    writefile.write(line+"\n")
     #               print(line)    #this is for testing only***
                    line=plaintiff+" vs. "+defs
                if i==last_row:    #print last entry if about to jump out of loop
                    line=findreplace(line)
                    writefile.write(line+"\n")
     #               print(line)    #this is for testing only***
                oldplaint=plaintiff
                i+=1
            wb.save(filename)
        except FileNotFoundError:
            print("No file: "+filename)
        if filename=='JUPD.xlsx':
            filename=''
        if filename=='CLPD.xlsx':
            filename='JUPD.xlsx'
        if filename=='DSJU.xlsx':
            filename='CLPD.xlsx'
        if filename=='DSPL.xlsx':
            filename='DSJU.xlsx'
        if filename=='DSLP.xlsx':
            filename='DSPL.xlsx'
          
    writefile.close()


# transfers.py --- original body
def transfers():
    # start of program, open xlsx file, create text file, initialize variables
    fn=load_file()
    wb=load_workbook(filename=fn)
    sh=wb.active
    textfile=open("transfers.txt","w")
    textfile.write("Real estate transfers\nThe following property transfers were recently recorded in the Logan County Auditor's Office:\n")
    old_convey="nil"
    old_loc="nil"
    old_loc2="nil"


    #main loop
    START=2 # this variable ignores the header line
    i=START
    last_row=sh.max_row
    while i<=last_row:

        convey_no="C"+str(i)
        convey=sh[convey_no].value

        if convey==old_convey: #this initial test tries to handle properties in more than one taxing district
               #it should handle multiple parcels in two taxing districts, but three or more may start producing irregular results
            loc=str(sh["E"+str(i)].value)[0:2]
            if loc!=old_loc and loc!=old_loc2:
                new_loc=tax_code(loc)
                location=location+", "+new_loc
                old_loc2=loc
            i+=1

        #This is the main body of the loop
        else:
            #print line if it's not the first line
            if i>START:
                line=date+": "+grantor+" to "+grantee+", property in "+location+", "+pricestr+"\n"
                line=check_line(line)
                textfile.write(line)

            #reset all values
            date=""
            location=""
            grantor=""
            grantee=""
            price=""

            #identifying date cell and setting date to AP style
            date_cell=sh["A"+str(i)].value
         #   date_data=datetime.strptime(date_cell[0:10], "%m/%d/%Y")
            date=AP_Style_date(date_cell)

            #discovering the location of the property from first two digits of its poperty ID
            loc_cell=sh["E"+str(i)].value
            loc=str(loc_cell[0:2])
            location=tax_code(loc)
            old_loc=loc

            #Grantor (columns I)
            grantor=sh["I"+str(i)].value
            if grantor==None: #this loop handles a hypotehtical case inferred from
                              #a grantee case discovered in a file
                if sh["C"+str(i+1)].value==convey and sh["I"+str(i+1)].value!=None:
                    grantor=sh["I"+str(i+1)].value
                elif sh["C"+str(i+2)].value==convey and sh["I"+str(i+2)].value!=None:
                    grantor=sh["I"+str(i+2)].value
            grantor=grantor.title()
            grantor=grantor_grantee(grantor)

            #Grantee (column J)
            grantee=sh["J"+str(i)].value
            if grantee=="": #This case was discovered in a file while writing the code
                print ("true")
                if sh["C"+str(i+1)].value==convey and sh["J"+str(i+1)].value!=None:
                    grantee=sh["J"+str(i+1)].value
                elif sh["C"+str(i+2)].value==convey and sh["J"+str(i+2)].value!=None:
                    grantee=sh["J"+str(i+2)].value
            grantee=grantee.title()
            grantee=grantor_grantee(grantee)

            
            #the sales price is in column K
            price=sh["K"+str(i)].value
            if price==0:
                pricestr="county fee exempt"
            else:
                pricestr="$"+str(format(price, ','))

            old_convey=convey
            i+=1

    line=date+": "+grantor+" to "+grantee+", property in "+location+", "+pricestr+"\n"
    line=check_line(line)
    textfile.write(line)
    textfile.close()

# CCP_civil.py -- original body
def CCP_civil():
    #open files
    fn=load_file()
    readfile=open(fn,"r")
    writefile=open("civilcases.txt", "w")
    writefile.write("Civil filings\nThe following civil cases were recently filed in Logan County Common Court:\n")


    #main loop
    plaintiff=""
    defendant=""
    oldcase=""
    filedate=""
    resetp=True
    resetd=True
    for line in readfile:
        data=line.split("\t")
        case=data[0][0:13]
        if case!=oldcase:
            if oldcase!="":
                dataline=filedate+": "+plaintiff+" vs. "+defendant+"; "+casetype+"\n"
                dataline=findreplace(dataline)
                writefile.write(dataline)
                resetp=True
                plaintiff=""
                resetd=True
                defendant=""
            oldcase=case
            fdate=datetime.strptime(data[2][0:10], "%m/%d/%Y")
            filedate=AP_Style_date(fdate)
            casetype=data[3].lower()
        party=data[5][0:9]
        party_name=str(data[4].title())
        while party_name[len(party_name)-1]==" ":
            party_name=party_name[0:len(party_name)-1]
        while party_name[0]==" ":
            party_name=party_name[1:len(party_name)]    
        party_name=name_field(party_name)
        if party=="PLAINTIFF" and resetp==False:
            plaintiff=plaintiff+", "+party_name
        elif party=="PLAINTIFF" and resetp==True:
            plaintiff=party_name
            resetp=False
        elif party=="DEFENDANT" and resetd==False:
            defendant=defendant+", "+party_name
        elif party=="DEFENDANT" and resetd==True:
            defendant=party_name
            resetd=False

    writefile.write(filedate+": "+plaintiff+" vs. "+defendant+"; "+casetype+"\n")
    readfile.close()
    writefile.close()


# fair_OC.py (this calls no functions)
#open files
def fair_OC():
    fn=load_file()
    f1=open(fn,"r")
    f2=open(cl+".txt", "w")
    cl=""
    cat=""
    gr=""
    first=""
    second=""
    third=""
    group=""

    for line in f1:
        if line!=None and line!="" and line!=" " and line!="\n":
            lin=str(line[0:len(line)-1])
            lin=lin.replace("  ", " ")
            lin=lin.replace("  ", " ")
            lin=lin.replace("  ", " ")
            lin=lin.replace("  ", " ")
            lin=lin.replace("  ", " ")
            lin=lin.replace("  ", " ")
            lin=lin.replace("  ", " ")
            lin=lin.replace("  ", " ")
            lin=lin.replace("  ", " ")
            lin=lin.replace("  ", " ")
            lin=lin.replace(" 1st -", "1.")
            lin=lin.replace(" 2nd -", "2.")
            lin=lin.replace(" 3rd -", "3.")
            if lin[0]=="1":
                lin=lin.title()
                first=lin
            elif lin[0]=="2":
                lin=lin.title()
                second=lin
            elif lin[0]=="3":
                lin=lin.title()
                third=lin            
            else:
                # print old line to file
                if second=="" and third=="":
                    f2.write(group+": "+first+"\n")
                elif third=="":
                    f2.write(group+": "+first+"; "+second+"\n")
                else:
                    f2.write(group+": "+first+"; "+second+"; "+third+"\n")
                first=""
                second=""
                third=""
                
                # start new category
                lin=lin.capitalize()
                data=lin.split(" - ")
                try:
                    if data[0]!=cl:
                        cl=data[0]
                        f2.write(cl+"\n")
                except IndexError:
                    data[0]=""
                try:
                    if data[1]!=cat:
                        cat=data[1]
                        f2.write(cat.capitalize()+"\n")
                except IndexError:
                    data[1]=""
                group=data[2].capitalize()

    f1.close()
    f2.close()



window=GUI()

